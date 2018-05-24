import csv
import pandas as pd
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import load_workbook, Workbook
import os.path

def flatten_json( b, delim ):
    val = {}
    for i in b.keys():
        if isinstance( b[i], dict ):
            get = flatten_json( b[i], delim )
            for j in get.keys():
                val[ i + delim + j ] = get[j]
        else:
            val[i] = b[i]

    return val

def data_to_gsheets(csv_filename, locations, scope, sheet_id):
    creds = ServiceAccountCredentials.from_json_keyfile_name('client_secret.json', scope)
    client = gspread.authorize(creds)
    sheet = client.open_by_key(sheet_id).sheet1
    # headers = sheet.row_values(1)
    # row = 2 # start from the second row because the first row are headers
    # modified_cells = []
    # for k in range(len(locations)):
    #     values=[]
    #     cell_list = sheet.range('A%s:K%s' % (row,row)) # make sure your row range equals the length of the values list
    #     for key in headers:
    #         if key in locations[k]:
    #             values.append(locations[k][key])
    #         else:
    #             values.append('')
    #     for i in range(len(cell_list)):
    #         cell_list[i].value = values[i]
    #     modified_cells += cell_list
    #     print("Updating row " + str(k+2) + '/' + str(len(locations) + 1))
    #     row += 1
    # sheet.update_cells(modified_cells)

    df = pd.DataFrame.from_records(locations)
    cols = list(df)
    cols.insert(0, cols.pop(cols.index('phone')))
    cols.insert(0, cols.pop(cols.index('name')))
    df = df.ix[:, cols]
    print(df)
    df.to_csv(csv_filename, encoding='utf-8', index=False)
    with open(csv_filename) as f:
        csv_file = f.read()
    print(csv_file)
    client.import_csv(sheet_id, csv_file)


def data_to_excel(locations, filename, sheetname):
    df = pd.DataFrame.from_records(locations)
    cols = list(df)
    cols.insert(0, cols.pop(cols.index('phone')))
    cols.insert(0, cols.pop(cols.index('name')))
    df = df.ix[:, cols]
    print(df)

    if not os.path.isfile(filename):
        wb = Workbook()
        wb.save(filename)

    book = load_workbook(filename)
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    writer.book = book
    df = df.applymap(lambda x: x.encode('unicode_escape').
                                   decode('utf-8') if isinstance(x, str) else x)
    df.to_excel(writer, sheetname, index=False, encoding='utf-8')
    writer.save()